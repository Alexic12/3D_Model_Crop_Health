"""
Streamlit UI for Crop Health Visualization.

Key fixes for deployment:
- Responsive right sidebar (no fixed 400px margin). Uses padding-right with media queries.
- All figures use container width; Plotly margins trimmed.
- HTML component (interactive QGIS/Google map) is embedded with scrolling=True and inside a wrapper that can overflow horizontally if needed.
"""

from __future__ import annotations

import logging
import os

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
import streamlit.components.v1 as components
from scipy.stats import gaussian_kde

# Configuration import
from app.config.config import settings

# HPC imports
from app.data.data_processing import (
    process_uploaded_file,
    load_timeseries_data,
    invert_climate_file_rows,
    bulk_unzip_and_analyze_new_parallel,
    Prospectiva,  # noqa: F401  (kept for users that call it elsewhere)
    run_full_hpc_pipeline,
)
from app.data.ghg_capture import (
    process_ghg_data,
    create_risk_matrix_heatmap,
    create_lda_distribution_plot,
    create_management_matrix_heatmap,
    create_cost_benefit_chart,
)

# Visualization helpers
from app.ui.visualization import (
    create_2d_scatter_plot_ndvi,
    create_2d_scatter_plot_ndvi_interactive_qgis,
    create_3d_surface_plot,
    create_3d_simulation_plot_sea_keypoints,
    create_3d_simulation_plot_time_interpolation,
)

logger = logging.getLogger(__name__)


def create_ghg_analysis_excel(ghg_data, field_name=None):
    """
    Create a comprehensive Excel file with GHG risk analysis data for professional use.
    
    Args:
        ghg_data: GHG analysis dictionary containing all risk scenarios and matrices
        field_name: Optional field name for context
    
    Returns:
        BytesIO: Excel file as bytes buffer
    """
    print("🔍 DEBUG: Starting GHG Excel generation")
    print(f"🔍 DEBUG: ghg_data keys: {list(ghg_data.keys()) if ghg_data else 'None'}")
    
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    import pandas as pd
    import numpy as np
    
    def convert_to_excel_compatible(value):
        """Convert numpy types and arrays to Excel-compatible Python types."""
        # Check if it's a numpy scalar first (has dtype but no shape or shape is ())
        if hasattr(value, 'dtype') and (not hasattr(value, 'shape') or value.shape == ()):
            if value.dtype.kind in 'fc':  # float or complex
                return float(value)
            elif value.dtype.kind in 'iu':  # integer or unsigned
                return int(value)
            else:
                return str(value)
        # Check if it's a numpy array (has shape and it's not empty)
        elif hasattr(value, 'shape') and hasattr(value, 'tolist'):
            return [convert_to_excel_compatible(item) for item in value.tolist()]
        elif isinstance(value, (list, tuple)):
            return [convert_to_excel_compatible(item) for item in value]
        else:
            return value
    
    def convert_labels_to_excel(labels):
        """Convert labels (which might be numpy arrays) to list of strings for Excel."""
        converted = convert_to_excel_compatible(labels)
        
        # Handle nested arrays - flatten if we get a list of lists
        if isinstance(converted, list) and len(converted) == 1 and isinstance(converted[0], list):
            # This handles cases like [['Muy Pocos', 'Pocos', ...]] -> ['Muy Pocos', 'Pocos', ...]
            flattened = converted[0]
            return [str(item) for item in flattened]
        elif isinstance(converted, list):
            return [str(item) for item in converted]
        else:
            return [str(converted)]
    
    # Create workbook
    wb = Workbook()
    # Remove default sheet safely
    if wb.active:
        wb.remove(wb.active)
    
    # Define styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
    subheader_font = Font(bold=True, color="000000")
    subheader_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # 1. Executive Summary Sheet
    summary_ws = wb.create_sheet("📊 Resumen Ejecutivo")
    summary_ws.append(["ANÁLISIS DE RIESGO Y CAPTURA DE GEI"])
    summary_ws.append([""])
    summary_ws.append(["Campo:", field_name or "N/A"])
    summary_ws.append(["Fecha de Análisis:", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M")])
    summary_ws.append([""])
    
    # Key metrics summary
    if "results" in ghg_data and not ghg_data["results"].empty:
        baseline_loss = ghg_data["results"].loc["Baseline", "Media (USD)"] if "Baseline" in ghg_data["results"].index else 0
        best_scenario = ghg_data["results"]["Media (USD)"].idxmin()
        best_loss = ghg_data["results"].loc[best_scenario, "Media (USD)"]
        max_co2 = ghg_data["results"]["TCO2(Ton.)"].max()
        
        summary_ws.append(["MÉTRICAS CLAVE"])
        summary_ws.append(["Pérdida Baseline (USD):", f"${baseline_loss:,.0f}"])
        summary_ws.append(["Mejor Escenario:", best_scenario])
        summary_ws.append(["Pérdida Mejor Escenario (USD):", f"${best_loss:,.0f}"])
        summary_ws.append(["Reducción Potencial (%):", f"{((baseline_loss - best_loss) / baseline_loss * 100):.1f}%"])
        summary_ws.append(["Máxima Captura CO2 (Ton):", f"{max_co2:.3f}"])
    
    # 2. Risk Scenarios Sheet
    scenarios_ws = wb.create_sheet("📈 Escenarios de Riesgo")
    scenarios_ws.append(["ESCENARIOS DE GESTIÓN DE RIESGOS"])
    scenarios_ws.append([""])
    
    if "results" in ghg_data and not ghg_data["results"].empty:
        # Add dataframe to worksheet with proper type conversion
        for r in dataframe_to_rows(ghg_data["results"], index=True, header=True):
            # Convert numpy types to Python types
            clean_row = []
            for item in r:
                if hasattr(item, 'dtype'):  # numpy type
                    if item.dtype.kind in 'fc':  # float or complex
                        clean_row.append(float(item))
                    elif item.dtype.kind in 'iu':  # integer or unsigned
                        clean_row.append(int(item))
                    else:
                        clean_row.append(str(item))
                else:
                    clean_row.append(item)
            scenarios_ws.append(clean_row)
    
    # 3. Impact Matrix Sheet
    print("🔍 DEBUG: Processing Impact Matrix")
    if "impact_matrix" in ghg_data:
        print("🔍 DEBUG: Found impact_matrix in ghg_data")
        impact_ws = wb.create_sheet("🎯 Matriz de Impacto")
        impact_ws.append(["MATRIZ DE IMPACTO - FRECUENCIA VS SEVERIDAD"])
        impact_ws.append([""])
        
        # Convert numpy array to proper format
        impact_matrix = ghg_data["impact_matrix"]
        print(f"🔍 DEBUG: impact_matrix type: {type(impact_matrix)}")
        print(f"🔍 DEBUG: impact_matrix shape: {impact_matrix.shape if hasattr(impact_matrix, 'shape') else 'No shape'}")
        
        if hasattr(impact_matrix, 'shape'):  # It's a numpy array
            # Add frequency labels as column headers - convert numpy arrays to lists
            freq_labels = ghg_data.get("frequency_labels", [f"Freq_{i}" for i in range(impact_matrix.shape[1])])
            sev_labels = ghg_data.get("severity_labels", [f"Sev_{i}" for i in range(impact_matrix.shape[0])])
            
            # Convert numpy arrays to Excel-compatible format
            freq_labels = convert_labels_to_excel(freq_labels)
            sev_labels = convert_labels_to_excel(sev_labels)
            
            print(f"🔍 DEBUG: Converted labels - freq: {len(freq_labels)} items, sev: {len(sev_labels)} items")
            
            # Create matrix with labels
            header_row = ["Severidad \\ Frecuencia"] + freq_labels
            impact_ws.append(header_row)
            
            for i, sev_label in enumerate(sev_labels):
                # Ensure sev_label is properly converted 
                sev_label_str = convert_to_excel_compatible(sev_label)
                sev_label_str = sev_label_str if isinstance(sev_label_str, str) else str(sev_label_str)
                
                row = [sev_label_str] + [convert_to_excel_compatible(impact_matrix[i, j]) for j in range(impact_matrix.shape[1])]
                impact_ws.append(row)
            
            print("🔍 DEBUG: Impact matrix processed successfully")
    
    # 4. Management Matrices Sheet
    if "management_matrices" in ghg_data:
        mgmt_ws = wb.create_sheet("🛡️ Matrices de Gestión")
        mgmt_ws.append(["MATRICES DE ESTRATEGIAS DE GESTIÓN"])
        mgmt_ws.append([""])
        
        for strategy_name, matrix in ghg_data["management_matrices"].items():
            mgmt_ws.append([f"ESTRATEGIA: {str(strategy_name)}"])
            
            if hasattr(matrix, 'shape'):  # It's a numpy array
                freq_labels = ghg_data.get("frequency_labels", [f"Freq_{i}" for i in range(matrix.shape[1])])
                sev_labels = ghg_data.get("severity_labels", [f"Sev_{i}" for i in range(matrix.shape[0])])
                
                # Convert numpy arrays to Excel-compatible format
                freq_labels = convert_labels_to_excel(freq_labels)
                sev_labels = convert_labels_to_excel(sev_labels)
                
                header_row = ["Severidad \\ Frecuencia"] + freq_labels
                mgmt_ws.append(header_row)
                
                for i, sev_label in enumerate(sev_labels):
                    # Ensure sev_label is properly converted 
                    sev_label_str = convert_to_excel_compatible(sev_label)
                    sev_label_str = sev_label_str if isinstance(sev_label_str, str) else str(sev_label_str)
                    row = [sev_label_str] + [convert_to_excel_compatible(matrix[i, j]) for j in range(matrix.shape[1])]
                    mgmt_ws.append(row)
            else:
                # Fallback for non-numpy data
                mgmt_ws.append(["Datos no disponibles en formato de matriz"])
            
            mgmt_ws.append([""])  # Separator between strategies
    
    # 5. LDA Analysis Sheet
    if "lda_data" in ghg_data:
        lda_ws = wb.create_sheet("📉 Análisis LDA")
        lda_ws.append(["ANÁLISIS DE DISTRIBUCIÓN DE PÉRDIDAS (LDA)"])
        lda_ws.append([""])
        
        try:
            print("🔍 DEBUG: Processing LDA data")
            # Convert LDA data to DataFrame if it's not already
            if isinstance(ghg_data["lda_data"], pd.DataFrame):
                lda_df = ghg_data["lda_data"]
                print(f"🔍 DEBUG: LDA data is already a DataFrame: {lda_df.shape}")
            else:
                # If it's array data, create a structured DataFrame
                lda_data = ghg_data["lda_data"]
                print(f"🔍 DEBUG: LDA data type: {type(lda_data)}")
                
                if hasattr(lda_data, 'shape'):  # numpy array
                    print(f"🔍 DEBUG: LDA data shape: {lda_data.shape}")
                    mgr_labels = ghg_data.get("mgr_labels", [f"Scenario_{i}" for i in range(lda_data.shape[0])])
                    print(f"🔍 DEBUG: mgr_labels length: {len(mgr_labels)} vs data rows: {lda_data.shape[0]}")
                    
                    # Fix: Create proper row indices based on actual data shape
                    if len(mgr_labels) != lda_data.shape[0]:
                        print(f"🔍 DEBUG: Mismatch detected - creating proper row indices")
                        mgr_labels = [f"Row_{i}" for i in range(lda_data.shape[0])]
                    
                    # Create column names if needed
                    if lda_data.shape[1] > 1:
                        column_names = [f"Column_{i}" for i in range(lda_data.shape[1])]
                        lda_df = pd.DataFrame(lda_data, index=mgr_labels, columns=column_names)
                    else:
                        lda_df = pd.DataFrame(lda_data, index=mgr_labels)
                    
                elif isinstance(lda_data, (list, tuple)):
                    print(f"🔍 DEBUG: LDA data is list/tuple with length: {len(lda_data)}")
                    mgr_labels = ghg_data.get("mgr_labels", [f"Scenario_{i}" for i in range(len(lda_data))])
                    if len(mgr_labels) != len(lda_data):
                        mgr_labels = [f"Row_{i}" for i in range(len(lda_data))]
                    lda_df = pd.DataFrame(lda_data, index=mgr_labels)
                else:
                    lda_ws.append(["Datos LDA no disponibles en formato compatible"])
                    lda_df = None
            
            if lda_df is not None:
                print(f"🔍 DEBUG: Final LDA DataFrame shape: {lda_df.shape}")
                print(f"🔍 DEBUG: LDA DataFrame columns: {list(lda_df.columns)}")
                
                # Add LDA data to worksheet
                for r in dataframe_to_rows(lda_df, index=True, header=True):
                    # Convert numpy types to Python types
                    clean_row = []
                    for item in r:
                        if hasattr(item, 'dtype'):  # numpy type
                            if item.dtype.kind in 'fc':  # float or complex
                                clean_row.append(float(item))
                            elif item.dtype.kind in 'iu':  # integer or unsigned
                                clean_row.append(int(item))
                            else:
                                clean_row.append(str(item))
                        else:
                            clean_row.append(item)
                    lda_ws.append(clean_row)
                print("🔍 DEBUG: LDA data processed successfully")
        except Exception as e:
            print(f"🚨 ERROR: LDA processing failed: {e}")
            import traceback
            print(f"🚨 ERROR: LDA traceback: {traceback.format_exc()}")
            lda_ws.append([f"Error procesando datos LDA: {str(e)}"])
    
    # 6. Financial Analysis Sheet
    financial_ws = wb.create_sheet("💰 Análisis Financiero")
    financial_ws.append(["ANÁLISIS COSTO-BENEFICIO"])
    financial_ws.append([""])
    
    if "results" in ghg_data and not ghg_data["results"].empty:
        # Extract financial columns
        financial_cols = ["Media (USD)", "VCap. (USD)", "IngOp.(USD)", "TCO2(Ton.)"]
        available_cols = [col for col in financial_cols if col in ghg_data["results"].columns]
        
        if available_cols:
            financial_data = ghg_data["results"][available_cols].copy()
            for r in dataframe_to_rows(financial_data, index=True, header=True):
                # Convert numpy types to Python types
                clean_row = []
                for item in r:
                    if hasattr(item, 'dtype'):  # numpy type
                        if item.dtype.kind in 'fc':  # float or complex
                            clean_row.append(float(item))
                        elif item.dtype.kind in 'iu':  # integer or unsigned
                            clean_row.append(int(item))
                        else:
                            clean_row.append(str(item))
                    else:
                        clean_row.append(item)
                financial_ws.append(clean_row)
    
    # Apply styling to all sheets
    for ws in wb.worksheets:
        # Style the first row (main headers)
        if ws.max_row > 0:
            for cell in ws[1]:
                if cell.value:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        from openpyxl.utils import get_column_letter
        for col_num, column in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_num)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)  # Cap at 30 characters
            ws.column_dimensions[column_letter].width = adjusted_width
    
        # Save to BytesIO buffer
        print("🔍 DEBUG: Saving workbook to buffer")
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        print("🔍 DEBUG: Excel generation completed successfully")
        return buffer


def save_hpc_workbook_to_disk(hpc_data, indice, anio, field_name):
    """
    Persist the HPC prospective workbook next to QGIS/IDW/Espacial files so the
    mobile app can consume it.

    Path: assets/data/{field_name}/INFORME_{indice}_HPC_{anio}.xlsx
    Returns the absolute path on success or None on failure.
    """
    try:
        if not field_name:
            logger.warning("save_hpc_workbook_to_disk: missing field_name; skipping persist.")
            return None
        buffer = create_complete_excel_data(hpc_data, indice)
        out_dir = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
            "assets", "data", field_name,
        )
        os.makedirs(out_dir, exist_ok=True)
        out_path = os.path.join(out_dir, f"INFORME_{indice}_HPC_{anio}.xlsx")
        with open(out_path, "wb") as fh:
            fh.write(buffer.getvalue())
        logger.info(f"HPC workbook persisted => {out_path}")
        return out_path
    except Exception as exc:
        logger.exception(f"Failed to persist HPC workbook: {exc}")
        return None


def create_complete_excel_data(hpc_data, indice):
    """
    Create a comprehensive Excel file with all HPC data for all points and dates.
    
    Args:
        hpc_data: HPC data dictionary containing results for all points
        indice: The vegetation index name (e.g., NDVI)
    
    Returns:
        BytesIO: Excel file as bytes buffer
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Create workbook
    wb = Workbook()
    # Remove default sheet safely
    if wb.active:
        wb.remove(wb.active)
    
    results = hpc_data.get("results", [])
    data_source = "Simulados" if hpc_data.get("using_mock_data", False) else "Reales"
    
    # Create summary sheet
    summary_ws = wb.create_sheet("📋 Resumen")
    summary_ws.append(["Resumen del Análisis HPC"])
    summary_ws.append(["Índice de vegetación:", indice])
    summary_ws.append(["Fuente de datos:", f"Datos {data_source}"])
    summary_ws.append(["Total de puntos:", len(results)])
    summary_ws.append(["Meses analizados:", "12"])
    summary_ws.append([""])
    summary_ws.append(["Punto", "Lat", "Lon", "Disponible"])
    
    # Add point summary
    for result in results:
        point_idx = result.get("point_idx", 0)
        display_idx = point_idx + 1  # 1-based display
        # Extract coordinates from HPC result
        lat = result.get("lat", "N/A")
        lon = result.get("lon", "N/A")
        summary_ws.append([f"Punto {display_idx}", lat, lon, "✅"])
    
    # Create detailed sheet for each point
    for result in results:
        point_idx = result.get("point_idx", 0)
        display_idx = point_idx + 1  # 1-based display
        sheet_name = f"📍 Punto {display_idx}"
        
        # Create worksheet for this point
        ws = wb.create_sheet(sheet_name)
        
        # Add header information
        ws.append([f"Datos HPC - Punto {display_idx}"])
        ws.append([f"Índice: {indice}"])
        ws.append([f"Fuente: Datos {data_source}"])
        ws.append([""])
        
        # Get data arrays
        XLDA = result.get("XLDA", None)
        VC = result.get("VC", [])
        XInf = result.get("XInf", None)
        
        if XLDA is not None and XInf is not None:
            # Create monthly data table
            columns = [
                "Mes", "WD", "Max C", "Min C", "Viento (m/s)", 
                "Humedad (%)", "Precip. (mm)", f"{indice}", "Skewness",
                "%C1", "%C2", "%C3", "Mean (USD)", "75% (USD)", "OpVar-99% (USD)"
            ]
            
            # Add column headers
            ws.append(columns)
            
            # Add data for each month
            n_months = min(12, XInf.shape[0] if XInf.ndim > 0 else 12)
            for month_i in range(n_months):
                wd = VC[month_i] if month_i < len(VC) else f"WD_{month_i}"
                
                if XInf.ndim > 1:
                    row_data = [
                        f"Mes {month_i + 1}",
                        wd,
                        f"{XInf[month_i, 0]:.3f}" if XInf.shape[1] > 0 else "N/A",
                        f"{XInf[month_i, 1]:.3f}" if XInf.shape[1] > 1 else "N/A",
                        f"{XInf[month_i, 2]:.3f}" if XInf.shape[1] > 2 else "N/A",
                        f"{XInf[month_i, 3]:.3f}" if XInf.shape[1] > 3 else "N/A",
                        f"{XInf[month_i, 4]:.3f}" if XInf.shape[1] > 4 else "N/A",
                        f"{XInf[month_i, 8]:.3f}" if XInf.shape[1] > 8 else "N/A",
                        f"{XInf[month_i, 5]:.3f}" if XInf.shape[1] > 5 else "N/A",
                        f"{XInf[month_i, 6]:.3f}" if XInf.shape[1] > 6 else "N/A",
                        f"{XInf[month_i, 7]:.3f}" if XInf.shape[1] > 7 else "N/A",
                        f"{XInf[month_i, 8]:.3f}" if XInf.shape[1] > 8 else "N/A",
                        f"{XInf[month_i, 9]:.3f}" if XInf.shape[1] > 9 else "N/A",
                        f"{XInf[month_i, 10]:.3f}" if XInf.shape[1] > 10 else "N/A",
                        f"{XInf[month_i, 11]:.3f}" if XInf.shape[1] > 11 else "N/A",
                    ]
                else:
                    # Fallback for 1D or unexpected shapes
                    row_data = [f"Mes {month_i + 1}", wd] + ["N/A"] * (len(columns) - 2)
                
                ws.append(row_data)
        else:
            ws.append(["No hay datos disponibles para este punto"])
    
    # Style the workbook
    for ws in wb.worksheets:
        # Style headers
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
    
    # Save to BytesIO buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _responsive_css() -> None:
    """Inject responsive CSS with mobile-first design and accessibility features."""
    st.markdown(
        """
        <style>
          /* CSS Custom Properties for Design System */
          :root {
              --primary-color: #1f77b4;
              --secondary-color: #ff7f0e;
              --success-color: #2ca02c;
              --warning-color: #ff7f0e;
              --error-color: #d62728;
              --background-color: #ffffff;
              --surface-color: #f8f9fa;
              --text-primary: #212529;
              --text-secondary: #6c757d;
              --border-color: #dee2e6;
              --border-radius: 8px;
              --spacing-xs: 0.25rem;
              --spacing-sm: 0.5rem;
              --spacing-md: 1rem;
              --spacing-lg: 1.5rem;
              --spacing-xl: 2rem;
              --font-size-sm: 0.875rem;
              --font-size-base: 1rem;
              --font-size-lg: 1.125rem;
              --font-size-xl: 1.25rem;
              --line-height-base: 1.5;
              --transition-base: 0.15s ease-in-out;
          }

          /* Dark Mode Support */
          @media (prefers-color-scheme: dark) {
              :root {
                  --background-color: #121212;
                  --surface-color: #1e1e1e;
                  --text-primary: #ffffff;
                  --text-secondary: #b3b3b3;
                  --border-color: #333333;
              }
          }

          /* Mobile First Base Styles */
          .main .block-container {
              padding: var(--spacing-sm);
              max-width: 100%;
              margin: 0 auto;
          }

          /* Responsive Sidebar */
          [data-testid="stSidebar"] {
              background-color: var(--surface-color);
              border-right: 1px solid var(--border-color);
              transition: transform var(--transition-base);
              overflow-y: auto;
              overflow-x: hidden;
          }
          
          [data-testid="stSidebar"] > div {
              overflow-y: auto;
              overflow-x: hidden;
              height: 100vh;
          }
          
          /* Force sidebar to right on all screen sizes */
          [data-testid="stSidebar"] > div {
              border-left: 1px solid var(--border-color);
              border-right: none;
          }

          /* Mobile: Collapsible sidebar */
          @media (max-width: 768px) {
              [data-testid="stSidebar"] {
                  position: fixed;
                  top: 0;
                  left: -100%;
                  height: 100vh;
                  width: 280px;
                  z-index: 1000;
                  transform: translateX(-100%);
              }
              
              [data-testid="stSidebar"].sidebar-open {
                  transform: translateX(0);
              }
              
              .main .block-container {
                  padding: var(--spacing-sm);
              }
          }

          /* Tablet Styles */
          @media (min-width: 769px) and (max-width: 1024px) {
              [data-testid="stSidebar"] {
                  position: fixed !important;
                  right: 0 !important;
                  left: auto !important;
                  width: 280px;
                  height: 100vh;
                  top: 0;
                  z-index: 999;
              }
              
              .main .block-container {
                  padding: var(--spacing-md) !important;
                  margin-right: 320px !important;
                  margin-left: var(--spacing-md) !important;
                  max-width: calc(100vw - 360px) !important;
              }
              
              .main {
                  margin-right: 280px !important;
              }
          }

          /* Desktop Styles */
          @media (min-width: 1025px) {
              [data-testid="stSidebar"] {
                  position: fixed !important;
                  right: 0 !important;
                  left: auto !important;
                  top: 0;
                  height: 100vh;
                  width: 320px;
                  box-shadow: -2px 0 8px rgba(0,0,0,0.1);
                  z-index: 999;
              }
              
              .main .block-container {
                  padding: var(--spacing-lg) !important;
                  margin-right: 360px !important;
                  margin-left: var(--spacing-lg) !important;
                  max-width: calc(100vw - 400px) !important;
                  width: calc(100vw - 400px) !important;
              }
              
              .main {
                  margin-right: 320px !important;
              }
          }

          /* Large Desktop */
          @media (min-width: 1400px) {
              .main .block-container {
                  max-width: calc(1200px - 320px) !important;
                  margin-right: 360px !important;
                  margin-left: var(--spacing-lg) !important;
              }
          }
          
          /* Ensure all content respects sidebar */
          @media (min-width: 769px) {
              .stApp > div:first-child {
                  margin-right: 320px !important;
              }
              
              [data-testid="stAppViewContainer"] {
                  margin-right: 320px !important;
              }
              
              .main > div {
                  margin-right: 0 !important;
              }
          }

          /* Responsive Typography */
          h1, h2, h3, h4, h5, h6 {
              line-height: var(--line-height-base);
              margin-bottom: var(--spacing-md);
              color: var(--text-primary);
          }

          /* Responsive Components */
          .stApp iframe {
              width: 100% !important;
              border-radius: var(--border-radius);
          }

          .overflow-wrap {
              width: 100%;
              overflow: auto;
              border-radius: var(--border-radius);
          }

          /* Hide skip navigation links */
          .skip-link {
              display: none !important;
          }
          
          /* Keep Streamlit's top header transparent but interactive so the
             sidebar expand button (which lives inside it in newer versions)
             remains clickable. We previously hid it entirely, which removed
             the only way to reopen a collapsed sidebar without reloading. */
          [data-testid="stHeader"] {
              background: transparent !important;
              height: auto !important;
              min-height: 0 !important;
              pointer-events: none; /* let clicks pass through empty header area */
          }
          [data-testid="stHeader"] * {
              pointer-events: auto;
          }
          /* Hide just the decorative toolbar items, not the whole header.
             stStatusWidget is intentionally kept visible so users see when
             the app is running long background tasks. */
          [data-testid="stToolbar"],
          [data-testid="stDecoration"] {
              display: none !important;
          }

          /* Make Streamlit's running indicator more prominent. */
          [data-testid="stStatusWidget"] {
              position: fixed !important;
              top: 10px !important;
              right: 60px !important;
              z-index: 1001 !important;
              background: #fff8e1 !important;
              border: 1px solid #f0c36d !important;
              border-radius: 6px !important;
              padding: 4px 10px !important;
              box-shadow: 0 2px 8px rgba(0,0,0,0.18) !important;
              font-weight: 600 !important;
          }

          /* Keep the collapsed-sidebar reopen button visible and pinned to the
             top-right (since the sidebar itself is positioned on the right).
             Covers selectors used across Streamlit versions. */
          [data-testid="stSidebarCollapsedControl"],
          [data-testid="collapsedControl"],
          [data-testid="stExpandSidebarButton"],
          button[kind="header"][aria-label*="sidebar" i],
          button[aria-label*="Open sidebar" i],
          button[aria-label*="Expand sidebar" i] {
              display: flex !important;
              visibility: visible !important;
              opacity: 1 !important;
              position: fixed !important;
              top: 8px !important;
              right: 8px !important;
              left: auto !important;
              z-index: 1001 !important;
              background: var(--surface-color, #ffffff) !important;
              border: 1px solid var(--border-color, #e0e0e0) !important;
              border-radius: 6px !important;
              box-shadow: 0 2px 6px rgba(0,0,0,0.15) !important;
              padding: 4px !important;
              pointer-events: auto !important;
          }
          [data-testid="stSidebarCollapsedControl"] svg,
          [data-testid="collapsedControl"] svg,
          [data-testid="stExpandSidebarButton"] svg {
              transform: scaleX(-1); /* point chevron toward the right-side sidebar */
          }
          
          /* Accessibility Improvements */
          button:focus,
          .stSelectbox > div > div:focus,
          .stTextInput > div > div > input:focus {
              outline: 2px solid var(--primary-color);
              outline-offset: 2px;
          }

          /* High Contrast Mode */
          @media (prefers-contrast: high) {
              :root {
                  --border-color: #000000;
                  --text-primary: #000000;
              }
          }

          /* Reduced Motion */
          @media (prefers-reduced-motion: reduce) {
              * {
                  animation-duration: 0.01ms !important;
                  animation-iteration-count: 1 !important;
                  transition-duration: 0.01ms !important;
              }
          }

          /* Touch-friendly Interactive Elements */
          @media (hover: none) and (pointer: coarse) {
              button, .stSelectbox, .stTextInput {
                  min-height: 44px;
              }
          }

          /* Fix Dropdown Text Visibility - Comprehensive */
          .stSelectbox label {
              color: var(--text-primary) !important;
          }
          
          .stSelectbox > div > div {
              color: var(--text-primary) !important;
              background-color: var(--background-color) !important;
          }
          
          .stSelectbox > div > div > div {
              color: var(--text-primary) !important;
              background-color: var(--background-color) !important;
          }
          
          .stSelectbox select {
              color: var(--text-primary) !important;
              background-color: var(--background-color) !important;
          }
          
          .stSelectbox option {
              color: var(--text-primary) !important;
              background-color: var(--background-color) !important;
          }
          
          /* Target all possible selectbox elements */
          [data-testid="stSelectbox"] {
              color: var(--text-primary) !important;
          }
          
          [data-testid="stSelectbox"] > div {
              color: var(--text-primary) !important;
          }
          
          [data-testid="stSelectbox"] div {
              color: var(--text-primary) !important;
          }
          
          [data-testid="stSelectbox"] span {
              color: var(--text-primary) !important;
          }
          
          /* Force text color on all child elements */
          .stSelectbox * {
              color: var(--text-primary) !important;
          }
          
          /* Ensure dropdown arrow is visible */
          .stSelectbox > div > div::after {
              border-color: var(--text-primary) transparent transparent transparent !important;
          }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_ui() -> None:
    st.set_page_config(
        page_title="Visualización de Salud de Cultivos",
        page_icon="🌿",
        layout="wide",
        initial_sidebar_state="expanded",
        menu_items={
            'Get Help': 'https://docs.streamlit.io',
            'Report a bug': 'mailto:support@crophealth.com',
            'About': 'Plataforma de Visualización de Salud de Cultivos v2.0'
        }
    )
    _responsive_css()
    
    # Initialize accessibility features
    from app.ui.accessibility import initialize_accessibility
    initialize_accessibility()
    
    # Add viewport meta tag for mobile responsiveness
    st.markdown(
        '<meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">',
        unsafe_allow_html=True
    )
    
    # Initialize responsive layout
    from app.ui.responsive_components import ResponsiveLayout
    ResponsiveLayout.responsive_container()

    # Lock page zoom — prevent scroll-wheel zoom everywhere
    st.markdown(
        """
        <style>
            /* Prevent pinch-zoom on touch devices */
            html, body {
                touch-action: pan-x pan-y;
                -ms-content-zooming: none;
            }
            /* Plotly modebar stays clickable */
            .js-plotly-plot .plotly .modebar {
                pointer-events: auto;
            }
        </style>
        <script>
            (function() {
                // Helper: attach zoom-block listeners to a window/document
                function blockZoom(win, doc) {
                    try {
                        // Block Ctrl+scroll (browser zoom)
                        doc.addEventListener('wheel', function(e) {
                            if (e.ctrlKey || e.metaKey) {
                                e.preventDefault();
                                e.stopPropagation();
                                return false;
                            }
                        }, { passive: false, capture: true });

                        // Block Ctrl+Plus / Ctrl+Minus / Ctrl+0 keyboard zoom
                        doc.addEventListener('keydown', function(e) {
                            if ((e.ctrlKey || e.metaKey) && (e.key === '+' || e.key === '-' || e.key === '=' || e.key === '0')) {
                                e.preventDefault();
                                e.stopPropagation();
                                return false;
                            }
                        }, { capture: true });
                    } catch(err) {}
                }

                // Block on current window
                blockZoom(window, document);

                // Block on parent window (Streamlit shell)
                try {
                    if (window.parent && window.parent !== window) {
                        blockZoom(window.parent, window.parent.document);
                    }
                } catch(e) {}

                // Block on top window
                try {
                    if (window.top && window.top !== window) {
                        blockZoom(window.top, window.top.document);
                    }
                } catch(e) {}

                // Block inside child iframes (mpld3, plotly embeds)
                function blockIframeZoom() {
                    document.querySelectorAll('iframe').forEach(function(iframe) {
                        try {
                            if (iframe.contentDocument) {
                                blockZoom(iframe.contentWindow, iframe.contentDocument);
                            }
                        } catch(e) {}
                    });
                }

                // Disable Plotly scroll zoom and watch for new elements
                var observer = new MutationObserver(function() {
                    blockIframeZoom();
                    document.querySelectorAll('.js-plotly-plot').forEach(function(plot) {
                        if (plot._fullLayout) {
                            try { Plotly.relayout(plot, {'scene.dragmode': 'turntable'}); } catch(e) {}
                        }
                    });
                });
                observer.observe(document.body, { childList: true, subtree: true });

                // Run once on load
                blockIframeZoom();
            })();
        </script>
        """,
        unsafe_allow_html=True,
    )

    # Hidden iframe that injects zoom-block JS into the parent Streamlit shell
    components.html(
        """
        <script>
        (function() {
            // Target every frame level: current, parent, top
            var targets = [window, window.parent, window.top];
            targets.forEach(function(w) {
                try {
                    w.document.addEventListener('wheel', function(e) {
                        if (e.ctrlKey || e.metaKey) {
                            e.preventDefault();
                            e.stopImmediatePropagation();
                            return false;
                        }
                    }, { passive: false, capture: true });
                    w.document.addEventListener('keydown', function(e) {
                        if ((e.ctrlKey || e.metaKey) &&
                            (e.key==='+' || e.key==='-' || e.key==='=' || e.key==='0')) {
                            e.preventDefault();
                            e.stopImmediatePropagation();
                            return false;
                        }
                    }, { capture: true });
                } catch(err) {}
            });
        })();
        </script>
        """,
        height=0,
    )

    # Always-visible custom toggle button injected into the parent Streamlit
    # document. Works even when Streamlit's built-in expand button is hidden
    # or its selector changes between versions.
    components.html(
        """
        <script>
        (function() {
            try {
                var topDoc = window.top.document;
                if (topDoc.getElementById('custom-sidebar-toggle')) return;

                var btn = topDoc.createElement('button');
                btn.id = 'custom-sidebar-toggle';
                btn.title = 'Mostrar/ocultar barra lateral';
                btn.innerHTML = '☰';
                btn.style.cssText = [
                    'position:fixed',
                    'top:10px',
                    'right:10px',
                    'z-index:2147483647',
                    'width:40px',
                    'height:40px',
                    'border-radius:8px',
                    'border:1px solid #d0d0d0',
                    'background:#ffffff',
                    'color:#222',
                    'font-size:20px',
                    'font-weight:bold',
                    'cursor:pointer',
                    'box-shadow:0 2px 8px rgba(0,0,0,0.2)',
                    'display:flex',
                    'align-items:center',
                    'justify-content:center',
                    'padding:0'
                ].join(';');

                btn.addEventListener('click', function() {
                    var sb = topDoc.querySelector('[data-testid="stSidebar"]');
                    if (!sb) return;
                    // Try Streamlit's own buttons first (handles state correctly)
                    var clickTargets = [
                        '[data-testid="stSidebarCollapseButton"] button',
                        '[data-testid="stSidebarCollapsedControl"] button',
                        '[data-testid="stExpandSidebarButton"]',
                        '[data-testid="collapsedControl"] button',
                        '[aria-label*="sidebar" i]'
                    ];
                    for (var i = 0; i < clickTargets.length; i++) {
                        var el = topDoc.querySelector(clickTargets[i]);
                        if (el) { el.click(); return; }
                    }
                    // Fallback: toggle visibility via inline style
                    var hidden = sb.getAttribute('aria-expanded') === 'false'
                              || sb.style.display === 'none'
                              || sb.offsetWidth === 0;
                    sb.style.display = hidden ? '' : 'none';
                });

                topDoc.body.appendChild(btn);
            } catch (e) {
                console.warn('custom-sidebar-toggle inject failed', e);
            }
        })();
        </script>
        """,
        height=0,
    )

    st.title("Visualización de Salud de Cultivos")
    st.write("---")
    page_mode = st.radio("Seleccionar Página de Visualización", ["Visualización 3D", "Visualización Prospectiva", "Gestión de Riesgos"], index=0)
    st.write("---")

    # ---------------- SIDEBAR ----------------
    with st.sidebar:
        st.header("Configuraciones")

        # Get available fields from existing folders
        base_folder = "./upload_data"
        available_fields = []
        if os.path.exists(base_folder):
            for item in os.listdir(base_folder):
                item_path = os.path.join(base_folder, item)
                if os.path.isdir(item_path):
                    available_fields.append(item)
        
        # Field selection dropdown - check for auto-detected field
        auto_detected = st.session_state.get("auto_detected_field", None)
        current_field = st.session_state.get("current_field_name", None)
        
        # Use current field name if available
        if current_field:
            field_name = current_field
            if current_field not in available_fields:
                available_fields.append(current_field)
        elif auto_detected:
            field_name = auto_detected
            if auto_detected not in available_fields:
                available_fields.append(auto_detected)
        
        if available_fields:
            # Set default to current/auto-detected field if available
            default_idx = 0
            if current_field and current_field in available_fields:
                default_idx = available_fields.index(current_field)
            elif auto_detected and auto_detected in available_fields:
                default_idx = available_fields.index(auto_detected)
            
            selected_field = st.selectbox("Seleccionar Campo", available_fields + ["Crear Nuevo Campo"], index=default_idx)
            if selected_field == "Crear Nuevo Campo":
                field_name = st.text_input("Nombre del Nuevo Campo", value="")
            else:
                field_name = selected_field
        else:
            field_name = st.text_input("Nombre del Campo", value="Perimetro Prev")
        
        indice = st.text_input("Índice de Vegetación", value="NDVI")
        anio = st.text_input("Año", value="2024")
        st.write("---")

        # Google Maps API Key from configuration
        google_api_key = settings.GOOGLE_MAPS_API_KEY
        if not google_api_key:
            st.info("No se configuró GOOGLE_MAPS_API_KEY. El mapa base interactivo puede estar limitado.")

        st.subheader("Análisis Masivo de ZIP NDVI")
        uploaded_files = st.file_uploader(
            "Subir pares .zip (base + ColorMap) y archivo Excel de clima", 
            type=["zip", "xlsx", "xls"], 
            accept_multiple_files=True
        )
        
        import re
        
        if uploaded_files:
            # Separate ZIP files and Excel files
            zip_files = [f for f in uploaded_files if f.name.lower().endswith('.zip')]
            excel_files = [f for f in uploaded_files if f.name.lower().endswith(('.xlsx', '.xls'))]
            
            if zip_files:
                # Extract field name from first ZIP file using improved regex patterns
                first_zip = zip_files[0].name
            detected_field = None
            
            # Try multiple patterns to detect field name (matching data_processing.py)
            patterns = [
                # Pattern 1: "001. Campo_Luna_Roja_NDVI_31ene2022.zip"
                r'\d+\.\s*([^_\s]+(?:[_\s][^_\s]+)*?)_+NDVI',
                # Pattern 2: "001. perimetro__prev_NDVI_31ene2022.zip"
                r'\d+\.\s*([^_]+(?:__[^_]+)*?)_+NDVI',
                # Pattern 3: Generic field extraction before NDVI
                r'\d+\.\s*(.+?)_NDVI'
            ]
            
            for pattern in patterns:
                match = re.search(pattern, first_zip, re.IGNORECASE)
                if match:
                    raw_field = match.group(1)
                    # Clean up the field name - keep underscores for folder names
                    detected_field = raw_field.replace('__', '_').strip().title().replace(' ', '_')
                    break
            
            if detected_field:
                # Always use detected field name, override current selection
                field_name = detected_field
                st.info(f"Campo auto-detectado: {detected_field}. Archivos listos para procesar.")
                # Store in session state to update dropdown and files
                st.session_state["auto_detected_field"] = detected_field
                st.session_state["uploaded_zip_files"] = zip_files
                st.session_state["current_field_name"] = detected_field
                # Only rerun once to update dropdown, then set flag to prevent further reruns
                if not st.session_state.get("field_updated", False):
                    st.session_state["field_updated"] = True
                    st.rerun()
            else:
                st.session_state["uploaded_zip_files"] = zip_files
                st.info("Archivos listos para procesar. Haz clic en 'Ejecutar Análisis Masivo' para procesar.")
            
            # Handle Excel files (climate data)
            if excel_files:
                st.session_state["uploaded_excel_files"] = excel_files
                climate_names = [f.name for f in excel_files]
                st.info(f"Archivos de clima detectados: {', '.join(climate_names)}")

        if st.button("Ejecutar Análisis Masivo"):
            if not field_name:
                st.error("Por favor selecciona o ingresa un nombre de campo.")
            elif "uploaded_zip_files" not in st.session_state:
                st.error("No se subieron archivos ZIP. Por favor sube archivos primero.")
            else:
                # Use auto-detected field name if available
                if "current_field_name" in st.session_state:
                    field_name = st.session_state["current_field_name"]
                    st.info(f"Usando nombre de campo detectado: {field_name}")
                
                # Extract field name from ZIP files as backup using improved patterns
                zip_files = st.session_state["uploaded_zip_files"]
                if not field_name or field_name in ["NDVI_2024", "Create New Field"]:
                    first_zip = zip_files[0].name
                    detected_field = None
                    
                    # Try multiple patterns to detect field name
                    patterns = [
                        # Pattern 1: "001. Campo_Luna_Roja_NDVI_31ene2022.zip"
                        r'\d+\.\s*([^_\s]+(?:[_\s][^_\s]+)*?)_+NDVI',
                        # Pattern 2: "001. perimetro__prev_NDVI_31ene2022.zip"
                        r'\d+\.\s*([^_]+(?:__[^_]+)*?)_+NDVI',
                        # Pattern 3: Generic field extraction before NDVI
                        r'\d+\.\s*(.+?)_NDVI'
                    ]
                    
                    for pattern in patterns:
                        match = re.search(pattern, first_zip, re.IGNORECASE)
                        if match:
                            raw_field = match.group(1)
                            # Clean up the field name - keep underscores for folder names
                            detected_field = raw_field.replace('__', '_').strip().title().replace(' ', '_')
                            break
                    
                    if detected_field:
                        field_name = detected_field
                        st.info(f"Nombre de campo re-detectado del ZIP: {field_name}")
                
                # First, save the uploaded files
                field_folder = os.path.join(base_folder, field_name)
                subfolder = os.path.join(field_folder, indice, anio)
                os.makedirs(subfolder, exist_ok=True)
                
                # Save ZIP files
                for zf in zip_files:
                    outpath = os.path.join(subfolder, zf.name)
                    with open(outpath, "wb") as f:
                        f.write(zf.getbuffer())
                
                # Save Excel files (climate data)
                if "uploaded_excel_files" in st.session_state:
                    excel_files = st.session_state["uploaded_excel_files"]
                    for ef in excel_files:
                        # Save with standard climate naming convention
                        climate_name = f"Clima_{indice}_{anio}.xlsx"
                        outpath = os.path.join(subfolder, climate_name)
                        with open(outpath, "wb") as f:
                            f.write(ef.getbuffer())
                        st.success(f"Archivo de clima guardado como: {climate_name}")
                
                st.success(f"Todos los archivos guardados en {field_name}/{indice}/{anio}.")
                
                # Then process the files
                with st.spinner("⏳ Ejecutando análisis masivo... Esto puede tardar varios minutos."):
                    esp_xlsx, idw_xlsx, qgis_xlsx = bulk_unzip_and_analyze_new_parallel(
                        indice, anio, base_folder=field_folder
                    )
                if esp_xlsx and os.path.exists(esp_xlsx):
                    st.success(f"Espacial => {esp_xlsx}")
                if idw_xlsx and os.path.exists(idw_xlsx):
                    st.success(f"IDW => {idw_xlsx}")
                    st.session_state["processed_idw_path"] = idw_xlsx
                    st.session_state["current_field"] = field_name
                if qgis_xlsx and os.path.exists(qgis_xlsx):
                    st.success(f"QGIS => {qgis_xlsx}")
                
                # Clear uploaded files and auto-detected field after processing
                if "uploaded_zip_files" in st.session_state:
                    del st.session_state["uploaded_zip_files"]
                if "uploaded_excel_files" in st.session_state:
                    del st.session_state["uploaded_excel_files"]
                if "auto_detected_field" in st.session_state:
                    del st.session_state["auto_detected_field"]
                if "current_field_name" in st.session_state:
                    del st.session_state["current_field_name"]
                if "field_updated" in st.session_state:
                    del st.session_state["field_updated"]
                # Trigger rerun to refresh dropdown
                st.rerun()

        # Always show download buttons if files exist
        if field_name:
            output_dir = os.path.join("assets", "data", field_name)
        else:
            output_dir = os.path.join("assets", "data")
        esp_file = os.path.join(output_dir, f"INFORME_{indice}_Espacial_{anio}.xlsx")
        idw_file = os.path.join(output_dir, f"INFORME_{indice}_IDW_{anio}.xlsx")
        qgis_file = os.path.join(output_dir, f"INFORME_{indice}_QGIS_{anio}.xlsx")
        
        st.write("---")
        st.subheader("Descargar Archivos Procesados")
        if os.path.exists(esp_file):
            with open(esp_file, "rb") as f:
                st.download_button("Descargar Espacial", data=f, file_name=os.path.basename(esp_file), key="dl_esp")
        if os.path.exists(idw_file):
            with open(idw_file, "rb") as f:
                st.download_button("Descargar IDW", data=f, file_name=os.path.basename(idw_file), key="dl_idw")
        if os.path.exists(qgis_file):
            with open(qgis_file, "rb") as f:
                st.download_button("Descargar QGIS", data=f, file_name=os.path.basename(qgis_file), key="dl_qgis")

        with st.expander("Invertir Filas de Excel de Clima"):
            climate_file = st.file_uploader("Subir Excel de Clima", type=["xlsx", "xls"])
            if climate_file:
                df_inv = invert_climate_file_rows(climate_file)
                if df_inv is not None:
                    st.dataframe(df_inv.head(20))
                    # Important: to_excel returns bytes only if using a BytesIO; streamlit handles file-like objects best.
                    from io import BytesIO

                    buf = BytesIO()
                    df_inv.to_excel(buf, index=False)
                    buf.seek(0)
                    st.download_button(
                        "Descargar Excel Invertido",
                        data=buf,
                        file_name="clima_inverted.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

        with st.expander("Deslizadores Avanzados", expanded=False):
            grid_size = st.slider("Resolución de Cuadrícula", 5, 300, 50, step=5)
            z_scale = st.slider("Escala Z", 0.1, 2.0, 1.0, step=0.1)
            smoothness = st.slider("Suavidad de Superficie (Gaussiana)", 0.0, 10.0, 1.0, step=0.1)
            color_map = st.selectbox("Mapa de Color", ["viridis", "plasma", "inferno", "magma", "cividis"])
            steps_value = st.slider("Pasos de interpolación temporal", 1, 20, 10)
            st.write("---")

    # ---------------- MAIN PAGE CONTENT ----------------
    # Responsive navigation for mobile
    if st.session_state.get('viewport_width', 1200) < 768:
        with st.expander("📱 Menú de Navegación", expanded=False):
            page_mode = st.radio("Seleccionar Página", ["Visualización 3D", "Visualización Prospectiva", "Gestión de Riesgos"], horizontal=True)
    
    if page_mode == "Visualización 3D":

        # Check for an existing processed IDW path
        current_field = st.session_state.get("current_field", field_name)
        if current_field:
            idw_file = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                "assets",
                "data",
                current_field,
                f"INFORME_{indice}_IDW_{anio}.xlsx",
            )
        else:
            idw_file = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                "assets",
                "data",
                f"INFORME_{indice}_IDW_{anio}.xlsx",
            )
        processed_disabled = True
        processed_path = None
        if "processed_idw_path" in st.session_state and os.path.exists(st.session_state["processed_idw_path"]):
            processed_disabled = False
            processed_path = st.session_state["processed_idw_path"]
        elif os.path.exists(idw_file):
            processed_disabled = False
            processed_path = idw_file

        visualize_processed_btn = st.button("Visualizar Datos Procesados", disabled=processed_disabled)
        if visualize_processed_btn:
            st.session_state["show_processed_data"] = True

        show_proc = st.session_state.get("show_processed_data", False)

        if show_proc and (processed_path is not None) and os.path.exists(processed_path):
            data_sheets = load_timeseries_data(processed_path)
            if data_sheets:
                sheet_list = list(data_sheets.keys())
                chosen_sheet_processed = st.selectbox(
                    "Seleccionar hoja para datos procesados (3D estático y 2D)", sheet_list, key="processed_sheet_selector"
                )

                col1, col2 = st.columns([1, 1], gap="small")
                with col1:
                    if current_field:
                        qgis_file = os.path.join(
                            os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                            "assets",
                            "data",
                            current_field,
                            f"INFORME_{indice}_QGIS_{anio}.xlsx",
                        )
                    else:
                        qgis_file = os.path.join(
                            os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                            "assets",
                            "data",
                            f"INFORME_{indice}_QGIS_{anio}.xlsx",
                        )
                    if not os.path.exists(qgis_file):
                        st.error(f"Archivo QGIS no encontrado => {qgis_file}")
                    else:
                        try:
                            xls = pd.ExcelFile(qgis_file)
                            if chosen_sheet_processed not in xls.sheet_names:
                                st.error(f"Hoja '{chosen_sheet_processed}' no está en QGIS => {xls.sheet_names}")
                            else:
                                df_qgis = pd.read_excel(qgis_file, sheet_name=chosen_sheet_processed)
                                html_2d = create_2d_scatter_plot_ndvi_interactive_qgis(
                                    qgis_df=df_qgis,
                                    sheet_name=chosen_sheet_processed,
                                    google_api_key=google_api_key,
                                    margin_frac=0.05,
                                )

                                if html_2d:
                                    st.markdown('<div class="overflow-wrap">', unsafe_allow_html=True)
                                    components.html(html_2d, height=700, scrolling=True)  # ← no key
                                    st.markdown("</div>", unsafe_allow_html=True)
                                else:
                                    st.error("No se pudo crear el gráfico interactivo QGIS.")
                        except Exception as e:
                            st.error(f"Error leyendo QGIS => {e}")

                with col2:
                    # Show corresponding NDVI ColorMap image instead of 3D plot
                    image_path = None
                    if current_field:
                        upload_folder = os.path.join("upload_data", current_field, indice, anio)
                    else:
                        upload_folder = os.path.join("upload_data", f"{indice}_{anio}")
                    if os.path.exists(upload_folder):
                        for file in os.listdir(upload_folder):
                            if "ColorMap" in file and chosen_sheet_processed in file and file.lower().endswith('.tiff'):
                                image_path = os.path.join(upload_folder, file)
                                break
                    
                    if image_path and os.path.exists(image_path):
                        from PIL import Image
                        import base64
                        from io import BytesIO
                        
                        img = Image.open(image_path)
                        # Upscale small images so they display large and crisp
                        min_display_width = 800
                        if img.width < min_display_width:
                            scale = min_display_width // img.width + 1
                            img = img.resize(
                                (img.width * scale, img.height * scale),
                                Image.NEAREST
                            )
                        
                        # Render as inline HTML (no iframe) so it zooms with the page
                        buffer = BytesIO()
                        img.save(buffer, format='PNG')
                        img_b64 = base64.b64encode(buffer.getvalue()).decode()
                        
                        ndvi_html = f"""
                        <html><body style="margin:0; padding:0; overflow:hidden;">
                        <div style="display:flex; flex-direction:column; align-items:flex-start;
                                    justify-content:flex-start; width:100%;
                                    box-sizing:border-box; padding:70px 0 0 0;">
                            <img src="data:image/png;base64,{img_b64}"
                                 style="display:block; width:65%; height:auto;
                                        border-radius:8px;
                                        box-shadow:0 2px 8px rgba(0,0,0,0.15);" />
                            <p style="text-align:center; width:75%; margin-top:8px; color:#666;
                                      font-family:sans-serif; font-size:14px;">
                                NDVI ColorMap - {chosen_sheet_processed}
                            </p>
                        </div>
                        <script>
                        (function() {{
                            var targets = [document];
                            try {{ if (window.parent && window.parent.document) targets.push(window.parent.document); }} catch(e) {{}}
                            try {{ if (window.top && window.top.document) targets.push(window.top.document); }} catch(e) {{}}
                            targets.forEach(function(doc) {{
                                doc.addEventListener('wheel', function(e) {{
                                    if (e.ctrlKey || e.metaKey) {{
                                        e.preventDefault();
                                        e.stopImmediatePropagation();
                                        return false;
                                    }}
                                }}, {{ passive: false, capture: true }});
                                doc.addEventListener('keydown', function(e) {{
                                    if ((e.ctrlKey || e.metaKey) && (e.key==='+' || e.key==='-' || e.key==='=' || e.key==='0')) {{
                                        e.preventDefault();
                                        e.stopImmediatePropagation();
                                        return false;
                                    }}
                                }}, {{ capture: true }});
                            }});
                        }})();
                        </script>
                        </body></html>
                        """
                        components.html(ndvi_html, height=520, scrolling=False)
                    else:
                        # Fallback to 3D plot if no image found
                        lat_arr = data_sheets[chosen_sheet_processed]["lat"]
                        lon_arr = data_sheets[chosen_sheet_processed]["lon"]
                        ndvi_mat = data_sheets[chosen_sheet_processed]["ndvi"]

                        x_vals, y_vals, z_vals = [], [], []
                        for i, latv in enumerate(lat_arr):
                            for j, lonv in enumerate(lon_arr):
                                x_vals.append(lonv)
                                y_vals.append(latv)
                                z_vals.append(ndvi_mat[i, j])

                        df_3d = pd.DataFrame({"Longitud": x_vals, "Latitud": y_vals, "NDVI": z_vals})
                        fig_3d = create_3d_surface_plot(df_3d, grid_size, color_map, z_scale, smoothness)
                        if fig_3d:
                            fig_3d.update_layout(margin=dict(l=0, r=0, t=30, b=0), autosize=True)
                            st.plotly_chart(fig_3d, use_container_width=True, config={'scrollZoom': False})

                fig_time = create_3d_simulation_plot_time_interpolation(
                    data_sheets, grid_size, color_map, z_scale, smoothness, steps_value,
                    frame_duration=200
                )
                if fig_time:
                    fig_time.update_layout(
                        margin=dict(l=0, r=0, t=30, b=100),
                        paper_bgcolor='rgba(0,0,0,0)',
                        plot_bgcolor='rgba(0,0,0,0)',
                        autosize=True,
                        width=None,   # remove fixed width so it fills container
                        height=900,
                    )
                    st.markdown("#### Animación 3D de Series Temporales")
                    # Render via components.html with embedded JS speed slider
                    # so speed changes happen client-side (no Streamlit rerun / chart rebuild)
                    import plotly.io as pio
                    chart_html = pio.to_html(
                        fig_time, include_plotlyjs='cdn', full_html=False,
                        config={'scrollZoom': False}
                    )
                    # Extract frame names from figure for JS animate call
                    _frame_names_js = [f.name for f in fig_time.frames]
                    import json as _json
                    _frame_names_json = _json.dumps(_frame_names_js)
                    speed_slider_html = f"""
                    <style>
                        html, body {{
                            margin: 0; padding: 0;
                            background: transparent !important;
                            overflow: hidden;
                        }}
                        .speed-control {{
                            display: flex; align-items: center; justify-content: center;
                            gap: 14px; padding: 8px 0 4px 0; font-family: sans-serif;
                        }}
                        .speed-control label {{
                            color: #ccc; font-size: 14px; white-space: nowrap;
                        }}
                        .speed-control input[type=range] {{
                            width: 260px; accent-color: #4682B4; cursor: pointer;
                        }}
                        .speed-control .speed-val {{
                            color: #4682B4; font-weight: bold; font-size: 14px;
                            min-width: 55px; text-align: center;
                        }}
                        /* Make plotly chart fill width and center */
                        .js-plotly-plot,
                        .plotly-graph-div,
                        .plot-container,
                        .svg-container {{
                            margin: 0 auto !important;
                            width: 100% !important;
                            max-width: 100% !important;
                        }}
                    </style>
                    <div class="speed-control">
                        <label>🎚️ Velocidad:</label>
                        <span style="color:#aaa;font-size:12px;">Rápido</span>
                        <input type="range" id="jsSpeedSlider" min="30" max="800" value="200" step="10">
                        <span style="color:#aaa;font-size:12px;">Lento</span>
                        <span class="speed-val" id="jsSpeedVal">200 ms</span>
                    </div>
                    {chart_html}
                    <script>
                    (function() {{
                        var frameNames = {_frame_names_json};
                        var slider = document.getElementById('jsSpeedSlider');
                        var valLabel = document.getElementById('jsSpeedVal');
                        var plotDiv = document.querySelector('.plotly-graph-div');
                        var isPlaying = false;

                        // Detect play/pause from Plotly button clicks
                        if (plotDiv) {{
                            plotDiv.on('plotly_animated', function() {{ isPlaying = false; }});
                            plotDiv.on('plotly_animatingframe', function() {{ isPlaying = true; }});
                        }}

                        slider.addEventListener('input', function() {{
                            var ms = parseInt(this.value);
                            valLabel.textContent = ms + ' ms';
                            if (plotDiv && isPlaying) {{
                                // Re-issue animate from current frame with new speed
                                Plotly.animate(plotDiv, frameNames, {{
                                    frame: {{duration: ms, redraw: true}},
                                    transition: {{duration: 0}},
                                    fromcurrent: true,
                                    mode: 'immediate'
                                }});
                            }}
                        }});

                        // Also disable scroll zoom on the 3D scene
                        if (plotDiv) {{
                            plotDiv.addEventListener('wheel', function(e) {{ e.preventDefault(); }}, {{passive: false}});
                        }}

                        // Force chart to fill container width
                        if (plotDiv) {{
                            Plotly.relayout(plotDiv, {{autosize: true}});
                            window.addEventListener('resize', function() {{
                                Plotly.Plots.resize(plotDiv);
                            }});
                        }}
                    }})();
                    </script>
                    """
                    # Inject transparent background on the Streamlit iframe itself
                    components.html(speed_slider_html, height=1020, scrolling=False)
                    st.markdown(
                        '<style>iframe[title="streamlit_components.v1.components.html"]'
                        '{background:transparent !important;}</style>',
                        unsafe_allow_html=True
                    )
            else:
                # Single-sheet processed fallback
                df_single = process_uploaded_file(processed_path)
                if df_single is None:
                    st.error("No se pudieron analizar los datos procesados. Puede que no coincidan con el formato esperado.")
                elif all(col in df_single.columns for col in ("Longitud", "Latitud", "NDVI")):
                    lat_vals = df_single["Latitud"].values
                    lon_vals = df_single["Longitud"].values
                    ndvi_vals = df_single["NDVI"].values

                    fig_2d = create_2d_scatter_plot_ndvi(
                        lat_vals, lon_vals, ndvi_vals, sheet_name="ProcessedSingle", google_api_key=google_api_key
                    )
                    fig_3d_static = create_3d_surface_plot(
                        df_single.rename(columns={"Longitud": "Longitud", "Latitud": "Latitud", "NDVI": "NDVI"}),
                        grid_size=grid_size,
                        color_map=color_map,
                        z_scale=z_scale,
                        smoothness=smoothness,
                    )
                    if fig_3d_static:
                        fig_3d_static.update_layout(margin=dict(l=0, r=0, t=30, b=0), autosize=True)

                    fig_wave = create_3d_simulation_plot_sea_keypoints(
                        df_single,
                        grid_size=grid_size,
                        color_map=color_map,
                        z_scale=z_scale,
                        smoothness=smoothness,
                        key_frames=50,
                        key_points=10,
                        wave_amplitude_range=(0.05, 0.4),
                        wave_frequency_range=(0.05, 0.2),
                        wave_speed_range=(0.5, 2.0),
                        influence_sigma=0.05,
                        random_seed=42,
                    )

                    col1, col2 = st.columns([1.3, 1], gap="medium")
                    with col1:
                        if fig_2d:
                            st.pyplot(fig_2d, use_container_width=True)
                    with col2:
                        if fig_3d_static:
                            st.plotly_chart(fig_3d_static, use_container_width=True)

                    if fig_wave:
                        fig_wave.update_layout(margin=dict(l=0, r=0, t=30, b=0))
                        st.markdown("#### Animación 3D Basada en Ondas (Hoja Única Procesada)")
                        st.plotly_chart(fig_wave, use_container_width=True)
                else:
                    st.error("Los datos procesados no contienen las columnas u hojas requeridas.")

    elif page_mode == "Visualización Prospectiva":
        st.write("## Visualización Prospectiva")

        # 1) HPC Data Option: run or reuse
        if "hpc_data" not in st.session_state:
            st.warning("No se encontraron datos HPC en la sesión. Haz clic en el botón para ejecutar el pipeline HPC.")
            if st.button("Ejecutar Pipeline HPC"):
                try:
                    if field_name:
                        # Ensure field_name uses underscores for consistency
                        safe_field_name = field_name.replace(' ', '_')
                        field_base_folder = os.path.join("./upload_data", safe_field_name)
                    else:
                        field_base_folder = "./upload_data"
                    with st.spinner("⏳ Ejecutando pipeline HPC... Esto puede tardar varios minutos."):
                        hpc_data = run_full_hpc_pipeline(indice, anio, base_folder=field_base_folder)
                    if hpc_data is None:
                        st.error("El pipeline HPC devolvió None. Revisa los logs o rutas de archivos.")
                    else:
                        st.session_state["hpc_data"] = hpc_data
                        # Auto-persist prospective workbook for the mobile app
                        saved_path = save_hpc_workbook_to_disk(hpc_data, indice, anio, field_name)
                        if saved_path:
                            st.info(f"💾 Datos prospectivos guardados para móvil: `{os.path.basename(saved_path)}`")
                        else:
                            st.warning("⚠️ No se pudo guardar el archivo prospectivo para la versión móvil.")
                        # Show data source information
                        if hpc_data.get("using_mock_data", False):
                            st.warning("⚠️ Pipeline HPC completado usando **datos de clima SIMULADOS** (no se subió ningún archivo de clima real). Los resultados son solo para pruebas.")
                        else:
                            st.success("✅ Pipeline HPC completado usando **datos de clima REALES**. Los resultados se basan en mediciones meteorológicas reales.")
                except Exception as e:
                    st.error(f"Error ejecutando pipeline HPC => {e}")
        else:
            hpc_data = st.session_state["hpc_data"]
            # Show current data source status
            if hpc_data.get("using_mock_data", False):
                st.info("📊 Datos HPC cargados (usando **datos de clima SIMULADOS**). Abajo puedes visualizar los resultados de prueba.")
            else:
                st.info("📊 Datos HPC cargados (usando **datos de clima REALES**). Abajo puedes visualizar los resultados.")

        # 2) IDW/QGIS Visualization
        current_field = st.session_state.get("current_field", field_name)
        if current_field:
            idw_file = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                "assets",
                "data",
                current_field,
                f"INFORME_{indice}_IDW_{anio}.xlsx",
            )
        else:
            idw_file = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                "assets",
                "data",
                f"INFORME_{indice}_IDW_{anio}.xlsx",
            )

        processed_disabled = True
        processed_path = None
        if "processed_idw_path" in st.session_state and os.path.exists(st.session_state["processed_idw_path"]):
            processed_disabled = False
            processed_path = st.session_state["processed_idw_path"]
        elif os.path.exists(idw_file):
            processed_disabled = False
            processed_path = idw_file

        data_sheets = None
        if processed_path and os.path.exists(processed_path):
            data_sheets = load_timeseries_data(processed_path)

        if data_sheets:
            sheet_list = list(data_sheets.keys())
            chosen_sheet_processed = st.selectbox(
                "Seleccionar hoja para datos procesados (3D estático y 2D)", sheet_list, key="processed_sheet_selector"
            )

            # HPC data selection
            HPC_info = None
            if "hpc_data" in st.session_state:
                hpc_data = st.session_state["hpc_data"]
                results = hpc_data.get("results", [])
                point_labels = [f"(Punto={r['point_idx'] + 1})" for r in results]
                if point_labels:
                    chosen_point = st.selectbox("Seleccionar resultado de punto HPC", point_labels)
                    chosen_idx = point_labels.index(chosen_point)
                    HPC_info = results[chosen_idx]

            col1, col2 = st.columns([1.3, 1], gap="medium")
            with col1:
                if current_field:
                    qgis_file = os.path.join(
                        os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                        "assets",
                        "data",
                        current_field,
                        f"INFORME_{indice}_QGIS_{anio}.xlsx",
                    )
                else:
                    qgis_file = os.path.join(
                        os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                        "assets",
                        "data",
                        f"INFORME_{indice}_QGIS_{anio}.xlsx",
                    )
                if not os.path.exists(qgis_file):
                    st.error(f"Archivo QGIS no encontrado => {qgis_file}")
                else:
                    try:
                        xls = pd.ExcelFile(qgis_file)
                        if chosen_sheet_processed not in xls.sheet_names:
                            st.error(f"Hoja '{chosen_sheet_processed}' no está en QGIS => {xls.sheet_names}")
                        else:
                            df_qgis = pd.read_excel(qgis_file, sheet_name=chosen_sheet_processed)
                            html_2d = create_2d_scatter_plot_ndvi_interactive_qgis(
                                qgis_df=df_qgis,
                                sheet_name=chosen_sheet_processed,
                                google_api_key=google_api_key,
                                margin_frac=0.05,
                            )

                            if html_2d:
                                # wrapper lets it scroll horizontally if needed
                                st.markdown('<div class="overflow-wrap">', unsafe_allow_html=True)
                                components.html(html_2d, height=700, scrolling=True)  # height must match function's height_px
                                st.markdown("</div>", unsafe_allow_html=True)
                            else:
                                st.error("No se pudo crear el gráfico interactivo QGIS.")
                    except Exception as e:
                        st.error(f"Error leyendo QGIS => {e}")

            with col2:
                lat_arr = data_sheets[chosen_sheet_processed]["lat"]
                lon_arr = data_sheets[chosen_sheet_processed]["lon"]
                ndvi_mat = data_sheets[chosen_sheet_processed]["ndvi"]

                x_vals, y_vals, z_vals = [], [], []
                for i, latv in enumerate(lat_arr):
                    for j, lonv in enumerate(lon_arr):
                        x_vals.append(lonv)
                        y_vals.append(latv)
                        z_vals.append(ndvi_mat[i, j])

                if HPC_info is not None:
                    hpc_data = st.session_state["hpc_data"]
                    XLDA = HPC_info["XLDA"]  # shape [1000 x 12]
                    # VC = HPC_info["VC"]    # unused in the density plot below
                    # XInf2 = HPC_info["XInf"]

                    # Add data source indicator to section header
                    data_source = "Datos Simulados" if hpc_data.get("using_mock_data", False) else "Datos Reales"
                    st.markdown(f"## Evolución de Riesgo Mensual ({data_source})")

                    fig = go.Figure()
                    n_months = XLDA.shape[1]

                    # Plot the filled-area KDE for HPC data
                    for m in range(1, n_months):
                        month_data = XLDA[:, m]
                        kde = gaussian_kde(month_data)
                        x_range = np.linspace(month_data.min(), month_data.max(), 200)
                        density = kde(x_range)
                        
                        # Define which months are initially visible (2, 6, 8, 10, 12)
                        initially_visible_months = [2, 3, 6, 9, 12]
                        is_visible = (m + 1) in initially_visible_months
                        
                        fig.add_trace(
                            go.Scatter(
                                x=x_range,
                                y=density,
                                mode="lines",
                                fill="tozeroy",
                                name=f"Mes {m+1}",
                                hovertemplate="Mes: %{text}<br>Pérdida: %{x:.2f}<br>Densidad: %{y:.2f}",
                                text=[f"Mes {m+1}"] * len(x_range),
                                visible=is_visible,  # Set initial visibility
                            )
                        )

                    fig.update_layout(
                        xaxis_title="Pérdidas (USD/Mes-Zona)",
                        yaxis_title="Densidad",
                        showlegend=True,
                        margin=dict(l=0, r=0, t=30, b=0),
                    )
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info("No hay datos HPC cargados/seleccionados. Ejecuta el pipeline o elige un punto.")

            st.markdown("---")

            # 3) HPC Risk Distributions & Table from HPC Data
            if "hpc_data" in st.session_state and HPC_info is not None:
                hpc_data = st.session_state["hpc_data"]
                XLDA = HPC_info["XLDA"]
                VC = HPC_info["VC"]
                XInf2 = HPC_info["XInf"]

                columns = [
                    "WD",
                    "Max C",
                    "Min  C",
                    "Viento (m/s)",
                    "Humedad (%)",
                    "Precip. (mm)",
                    f"{indice}",
                    "Skewness",
                    "%C1",
                    "%C2",
                    "%C3",
                    "Mean (units)",
                    "75% (units)",
                    "OpVar-99% (units)",
                ]

                n_months = XLDA.shape[1] if XLDA is not None else 12
                table_data = []
                for row_i in range(n_months):
                    wd = VC[row_i] if row_i < len(VC) else ""
                    maxC = XInf2[row_i, 0]
                    minC = XInf2[row_i, 1]
                    viento = XInf2[row_i, 2]
                    hum = XInf2[row_i, 3]
                    prec = XInf2[row_i, 4]
                    ndvi_val = XInf2[row_i, 8]  # or 9 depending on your schema
                    skewv = XInf2[row_i, 5]
                    pc1 = XInf2[row_i, 6]
                    pc2 = XInf2[row_i, 7]
                    pc3 = XInf2[row_i, 8]
                    mean_val = XInf2[row_i, 9]
                    p75_val = XInf2[row_i, 10]
                    opv99 = XInf2[row_i, 11]

                    table_data.append(
                        [wd, maxC, minC, viento, hum, prec, ndvi_val, skewv, pc1, pc2, pc3, mean_val, p75_val, opv99]
                    )

                df_hpc = pd.DataFrame(table_data, columns=columns)
                for col_ in df_hpc.columns.drop("WD"):
                    df_hpc[col_] = df_hpc[col_].astype(float).map("{:.3f}".format)

                # Add data source indicator to table header
                data_source = "Datos Simulados" if hpc_data.get("using_mock_data", False) else "Datos Reales"
                st.markdown(f"### Tabla de Datos de Riesgo HPC ({data_source})")
                st.dataframe(df_hpc)
                
                # Add download button for all points and dates
                st.markdown("---")
                st.markdown("### 📥 Descarga de Datos Completos")
                
                col_download1, col_download2 = st.columns([1, 2])
                with col_download1:
                    if st.button("📊 Descargar Excel Completo", help="Descarga todos los puntos y fechas en un archivo Excel"):
                        try:
                            # Create Excel data for all points
                            excel_buffer = create_complete_excel_data(hpc_data, indice)
                            
                            st.download_button(
                                label="💾 Guardar archivo Excel",
                                data=excel_buffer,
                                file_name=f"HPC_Datos_Completos_{indice}_{data_source.replace(' ', '_')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                help="Descarga el archivo con todos los datos de todos los puntos y fechas"
                            )
                            st.success("✅ Archivo Excel preparado para descarga")
                            
                        except Exception as e:
                            st.error(f"Error generando archivo Excel: {e}")
                
                with col_download2:
                    total_points = len(hpc_data.get("results", []))
                    st.info(f"""
                    **Contenido del archivo Excel:**
                    - 📍 **{total_points} puntos** de análisis
                    - 📅 **12 meses** de predicción por punto  
                    - 📊 **Todas las variables** HPC (clima, riesgo, estadísticas)
                    - 🏷️ **Origen:** {data_source}
                    """)
                    
            else:
                st.info("No hay datos HPC cargados. Por favor haz clic en 'Ejecutar Pipeline HPC' para calcular resultados.")
        else:
            st.info("No hay datos IDW para visualizar o archivo no encontrado. Sube o procesa .zip para datos NDVI si es necesario.")

    elif page_mode == "Gestión de Riesgos":
        st.write("## Captura de GEI y Gestión de Riesgos")
        
        # Process GHG data
        if st.button("Procesar Datos de GEI"):
            with st.spinner("Procesando datos de captura de GEI..."):
                # Ensure field_name uses underscores for consistency
                safe_field_name = field_name.replace(' ', '_') if field_name else None
                ghg_data = process_ghg_data(indice, anio, base_folder="./upload_data", field_name=safe_field_name)
                if ghg_data:
                    st.session_state["ghg_data"] = ghg_data
                    st.success("¡Datos de GEI procesados exitosamente!")
                else:
                    st.error("Falló el procesamiento de datos de GEI. Verifica si existen los archivos requeridos.")
        
        # Display results if data exists
        if "ghg_data" in st.session_state:
            ghg_data = st.session_state["ghg_data"]
            
            # Display cluster information
            st.subheader("Análisis de Clusters")
            st.info(f"Tasa de Cambio: 1 USD = {ghg_data['usd_cop_rate']:,.0f} COP")
            st.dataframe(ghg_data["clusters"])
            
            # Risk matrices visualization
            col1, col2 = st.columns(2)
            
            with col1:
                st.subheader("Matriz de Eventos")
                fig_events = create_risk_matrix_heatmap(
                    ghg_data["events_matrix"], 
                    "Matriz de Eventos",
                    ghg_data["frequency_labels"],
                    ghg_data["severity_labels"]
                )
                st.plotly_chart(fig_events, use_container_width=True)
            
            with col2:
                st.subheader("Matriz de Pérdidas")
                fig_losses = create_risk_matrix_heatmap(
                    ghg_data["losses_matrix"], 
                    "Matriz de Pérdidas (USD)",
                    ghg_data["frequency_labels"],
                    ghg_data["severity_labels"]
                )
                st.plotly_chart(fig_losses, use_container_width=True)
            
            # Management results table
            st.subheader("Escenarios de Gestión de Riesgos")
            
            # Create tabs for USD and COP display
            tab1, tab2 = st.tabs(["💰 Valores en COP", "💵 Valores en USD"])
            
            with tab1:
                st.markdown("**Tabla en Pesos Colombianos (COP)**")
                if "results_cop" in ghg_data:
                    st.dataframe(ghg_data["results_cop"])
                else:
                    # Fallback if COP version not available
                    st.dataframe(ghg_data["results"])
                    st.info("💡 Conversión COP no disponible. Mostrando valores en USD.")
            
            with tab2:
                st.markdown("**Tabla en Dólares Americanos (USD)**")
                st.dataframe(ghg_data["results"])
                if "usd_cop_rate" in ghg_data:
                    st.caption(f"💱 Tasa de cambio: 1 USD = {ghg_data['usd_cop_rate']:,.0f} COP")
            
            # Combined matrices section: Impact matrix and Management matrices
            st.subheader("Matrices de Gestión")
            
            # Impact matrix
            st.markdown("### Matriz de Impacto")
            fig_impact = create_risk_matrix_heatmap(
                ghg_data["impact_matrix"], 
                "Matriz de Impacto",
                ghg_data["frequency_labels"],
                ghg_data["severity_labels"]
            )
            st.plotly_chart(fig_impact, use_container_width=True)
            
            # Management matrices visualization
            if "management_matrices" in ghg_data:
                st.markdown("### Matrices de Estrategias de Gestión")
                matrix_cols = st.columns(2)
                matrix_names = list(ghg_data["management_matrices"].keys())
                
                for i, (name, matrix) in enumerate(ghg_data["management_matrices"].items()):
                    col_idx = i % 2
                    with matrix_cols[col_idx]:
                        fig_matrix = create_management_matrix_heatmap(
                            matrix,
                            f"Matriz de Gestión - {name}",
                            ghg_data["frequency_labels"],
                            ghg_data["severity_labels"]
                        )
                        st.plotly_chart(fig_matrix, use_container_width=True)
            
            # Cost-benefit analysis
            st.subheader("Análisis Costo-Beneficio")
            
            # Create tabs for USD and COP cost-benefit analysis
            cb_tab1, cb_tab2 = st.tabs(["💰 Análisis en COP", "💵 Análisis en USD"])
            
            with cb_tab1:
                st.markdown("**Análisis Costo-Beneficio en Pesos Colombianos (COP)**")
                if "results_cop" in ghg_data:
                    fig_cost_benefit_cop = create_cost_benefit_chart(ghg_data["results_cop"])
                    st.plotly_chart(fig_cost_benefit_cop, use_container_width=True)
                else:
                    st.info("💡 Análisis COP no disponible. Ver pestaña USD.")
            
            with cb_tab2:
                st.markdown("**Análisis Costo-Beneficio en Dólares Americanos (USD)**")
                fig_cost_benefit_usd = create_cost_benefit_chart(ghg_data["results"])
                st.plotly_chart(fig_cost_benefit_usd, use_container_width=True)
            
            # LDA distribution plot
            st.subheader("Análisis de Distribución de Pérdidas (LDA)")
            
            # Create tabs for USD and COP LDA plots
            lda_tab1, lda_tab2 = st.tabs(["💰 Distribución en COP", "💵 Distribución en USD"])
            
            with lda_tab1:
                st.markdown("**Distribución de Pérdidas en Pesos Colombianos (COP)**")
                fig_lda_cop = create_lda_distribution_plot(
                    ghg_data["lda_data"],
                    ghg_data["mgr_labels"],
                    ghg_data.get("visualization_lines"),
                    currency="COP",
                    usd_cop_rate=ghg_data.get("usd_cop_rate", 1)
                )
                st.plotly_chart(fig_lda_cop, use_container_width=True)
            
            with lda_tab2:
                st.markdown("**Distribución de Pérdidas en Dólares Americanos (USD)**")
                fig_lda_usd = create_lda_distribution_plot(
                    ghg_data["lda_data"],
                    ghg_data["mgr_labels"],
                    ghg_data.get("visualization_lines"),
                    currency="USD"
                )
                st.plotly_chart(fig_lda_usd, use_container_width=True)
        
            # Comprehensive metrics dashboard
            st.subheader("Resumen de Métricas de Riesgo")
            
            # Create tabs for USD and COP metrics
            metrics_tab1, metrics_tab2 = st.tabs(["💰 Métricas en COP", "💵 Métricas en USD"])
            
            with metrics_tab1:
                st.markdown("**Métricas de Riesgo en Pesos Colombianos (COP)**")
                
                # Check if COP data is available
                if "results_cop" in ghg_data:
                    results_data = ghg_data["results_cop"]
                    currency_symbol = ""
                    currency_format = "{:,.0f} COP"
                    
                    # Get COP conversion rate for visualization lines
                    usd_cop_rate = ghg_data.get("usd_cop_rate", 1)
                    
                    # Basic risk metrics
                    st.markdown("### Métricas Básicas de Riesgo")
                    metrics_cols = st.columns(4)
                    
                    with metrics_cols[0]:
                        baseline_loss = results_data.loc["Baseline", "Media (COP)"]
                        st.metric("Pérdida Base", currency_format.format(baseline_loss))
                    
                    with metrics_cols[1]:
                        best_scenario = results_data["Media (COP)"].idxmin()
                        best_loss = results_data.loc[best_scenario, "Media (COP)"]
                        reduction = ((baseline_loss - best_loss) / baseline_loss) * 100
                        st.metric("Mejor Escenario", best_scenario, f"-{reduction:.1f}%")
                    
                    with metrics_cols[2]:
                        max_vc = results_data["VCap. (COP)"].max()
                        st.metric("Máximo Valor Capturado", currency_format.format(max_vc))
                    
                    with metrics_cols[3]:
                        total_events = results_data["NE"].iloc[0]
                        st.metric("Total Eventos de Riesgo", f"{total_events:,}")

                    # Advanced risk analysis
                    st.markdown("### Análisis Avanzado de Riesgo")
                    enhanced_cols = st.columns(4)
                    
                    with enhanced_cols[0]:
                        # CO2 Capture metrics
                        max_co2 = results_data["TCO2(Ton.)"].max()
                        st.metric("Máxima Captura CO2", f"{max_co2:.3f} toneladas")
                    
                    with enhanced_cols[1]:
                        # Skewness analysis
                        baseline_skew = results_data.loc["Baseline", "C.As."]
                        st.metric("Asimetría Base", f"{baseline_skew:.4f}")
                    
                    with enhanced_cols[2]:
                        # Operational Income
                        max_op_income = results_data["IngOp.(COP)"].max()
                        st.metric("Máximo Ingreso Op.", currency_format.format(max_op_income))
                    
                    with enhanced_cols[3]:
                        # Reference lines info
                        if "visualization_lines" in ghg_data:
                            media_improvement = ((ghg_data["visualization_lines"]["media_val_o"] - 
                                               ghg_data["visualization_lines"]["media_val_g"]) * usd_cop_rate)
                            st.metric("Reducción de Riesgo", currency_format.format(media_improvement))
                        
                else:
                    st.info("💡 Métricas COP no disponibles. Ver pestaña USD.")
            
            with metrics_tab2:
                st.markdown("**Métricas de Riesgo en Dólares Americanos (USD)**")
                
                results_data = ghg_data["results"]
                currency_format = "${:,.0f}"
                
                # Basic risk metrics
                st.markdown("### Métricas Básicas de Riesgo")
                metrics_cols = st.columns(4)
                
                with metrics_cols[0]:
                    baseline_loss = results_data.loc["Baseline", "Media (USD)"]
                    st.metric("Pérdida Base", currency_format.format(baseline_loss))
                
                with metrics_cols[1]:
                    best_scenario = results_data["Media (USD)"].idxmin()
                    best_loss = results_data.loc[best_scenario, "Media (USD)"]
                    reduction = ((baseline_loss - best_loss) / baseline_loss) * 100
                    st.metric("Mejor Escenario", best_scenario, f"-{reduction:.1f}%")
                
                with metrics_cols[2]:
                    max_vc = results_data["VCap. (USD)"].max()
                    st.metric("Máximo Valor Capturado", currency_format.format(max_vc))
                
                with metrics_cols[3]:
                    total_events = results_data["NE"].iloc[0]
                    st.metric("Total Eventos de Riesgo", f"{total_events:,}")

                # Advanced risk analysis
                st.markdown("### Análisis Avanzado de Riesgo")
                enhanced_cols = st.columns(4)
                
                with enhanced_cols[0]:
                    # CO2 Capture metrics
                    max_co2 = results_data["TCO2(Ton.)"].max()
                    st.metric("Máxima Captura CO2", f"{max_co2:.3f} toneladas")
                
                with enhanced_cols[1]:
                    # Skewness analysis
                    baseline_skew = results_data.loc["Baseline", "C.As."]
                    st.metric("Asimetría Base", f"{baseline_skew:.4f}")
                
                with enhanced_cols[2]:
                    # Operational Income
                    max_op_income = results_data["IngOp.(USD)"].max()
                    st.metric("Máximo Ingreso Op.", currency_format.format(max_op_income))
                
                with enhanced_cols[3]:
                    # Reference lines info
                    if "visualization_lines" in ghg_data:
                        media_improvement = (ghg_data["visualization_lines"]["media_val_o"] - 
                                           ghg_data["visualization_lines"]["media_val_g"])
                        st.metric("Reducción de Riesgo", currency_format.format(media_improvement))
            
            # Professional Excel export section
            st.markdown("---")
            st.markdown("### 📊 Exportación de Datos")
            
            col_download1, col_download2 = st.columns([1, 2])
            with col_download1:
                if st.button("📈 Descargar Análisis Completo", help="Descarga el reporte completo de análisis de riesgos en formato Excel"):
                    try:
                        # Create comprehensive Excel report
                        current_field = st.session_state.get("current_field", field_name)
                        excel_buffer = create_ghg_analysis_excel(ghg_data, current_field)
                        
                        # Generate professional filename
                        timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M")
                        field_suffix = f"_{current_field}" if current_field else ""
                        filename = f"Analisis_Riesgos_GEI{field_suffix}_{timestamp}.xlsx"
                        
                        st.download_button(
                            label="💾 Guardar Reporte Excel",
                            data=excel_buffer,
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            help="Descarga el análisis completo de riesgos y captura de GEI"
                        )
                        st.success("✅ Reporte Excel preparado para descarga")
                        
                    except Exception as e:
                        print(f"🚨 ERROR: Exception in Excel generation: {type(e).__name__}: {e}")
                        import traceback
                        print(f"🚨 ERROR: Full traceback: {traceback.format_exc()}")
                        st.error(f"Error generando reporte Excel: {e}")
                        st.error(f"Tipo de error: {type(e).__name__}")
                        st.error("Revisa la consola para más detalles del error")
            
            with col_download2:
                st.info("""
                **📋 Contenido del Reporte Excel:**
                - 📊 **Resumen Ejecutivo** con métricas clave
                - 📈 **Escenarios de Riesgo** detallados  
                - 🎯 **Matriz de Impacto** (frecuencia vs severidad)
                - 🛡️ **Matrices de Gestión** por estrategia
                - 📉 **Análisis LDA** (distribución de pérdidas)
                - 💰 **Análisis Financiero** y costo-beneficio
                """)
        
        else:
            st.info("Haz clic en 'Procesar Datos de GEI' para analizar escenarios de captura de gases de efecto invernadero y gestión de riesgos.")


# ---------------------- MAIN EXECUTION ENTRY ----------------------
if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    render_ui()
